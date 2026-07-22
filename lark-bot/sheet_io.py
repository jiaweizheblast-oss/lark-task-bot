"""
通用表格引擎：一套「生成 / 解析 / 防重复」通吃两种 HR 表。

- 列定义参数化（column spec）→ 同一个 build_xlsx / parse_rows 引擎驱动：
    · 渠道汇总表（第二个程序，现有）：一行一个渠道，数字型
    · 候选人联系表（第一个程序，爬取数据）：一行一个候选人，带「联系状态」下拉
- 网站和 Lark 机器人都调这里，数据天然互通。
- 本模块只负责「表格机制」：生成、解析、字段约束、防重复上传。
  它不写候选人真相层（Candidate / ContactActivity 等由核心 service 负责）——
  候选人表的解析只产出结构化行，交给上层，不在这里落库。

公共 API（保持向后兼容，bot.py 现有调用不用改）：
    build_template_xlsx(jobs, day, by)                          -> 渠道空表字节
    parse_sheet(data, filename, jobs, default_by, default_date) -> {rows, skipped, errors}
新增（候选人表，机制就绪，等第一个程序的候选人数据接入）：
    build_candidate_template_xlsx(jobs, day, candidates, sources)
    parse_candidate_sheet(data, filename, jobs, sources)        -> {rows, skipped, errors}
"""
import csv
import base64
import hashlib
import hmac
import io
import json
import uuid
import zipfile
from datetime import date
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from channel_report import CHANNELS, PIPELINE_STATUS
import channel_pipeline_schema as pipeline_schema


WORKBOOK_META_SHEET = "_NEXUS_META"
WORKBOOK_ARTIFACT_TYPE = "channel-candidate-pipeline"
WORKBOOK_META_VERSION = "channel-workbook-v3"
MINIMUM_SIGNING_KEY_BYTES = 32
MAX_XLSX_COMPRESSED_BYTES = 16 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 32 * 1024 * 1024
MAX_XLSX_MEMBERS = 250


def _validate_xlsx_container(data):
    if len(data) > MAX_XLSX_COMPRESSED_BYTES:
        raise ValueError("The workbook exceeds the 16 MB safety limit.")
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise ValueError("The workbook contains too many internal files.")
            total = 0
            for member in members:
                name = member.filename.replace("\\", "/").casefold()
                if member.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise ValueError("The workbook contains an oversized internal file.")
                total += member.file_size
                if total > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise ValueError("The workbook expands beyond the 64 MB safety limit.")
                if name.startswith("xl/externallinks/") or name.endswith("vbaproject.bin"):
                    raise ValueError("External links and macros are not accepted in HR workbooks.")
    except zipfile.BadZipFile as exc:
        raise ValueError("The uploaded file is not a valid XLSX workbook.") from exc


def _safe_excel_value(value, kind):
    if kind == "text" and isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _workbook_signature(metadata, signing_key):
    key = str(signing_key or "").encode("utf-8")
    if len(key) < MINIMUM_SIGNING_KEY_BYTES:
        raise ValueError("NEXUS_INTEGRATION_SIGNING_KEY is not configured safely")
    canonical = json.dumps(
        metadata, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hmac.new(key, b"channel-workbook-v2\n" + canonical, hashlib.sha256).hexdigest()


def _catalog_snapshot(jobs):
    return [
        {
            "job_ref": str(job.get("job_ref") or ("legacy-id-%s" % job.get("id"))),
            "title": str(job.get("title") or ""),
            "catalog_revision": int(job.get("catalog_revision") or 1),
            "accept_new": str(job.get("status") or "open").casefold() == "open",
        }
        for job in sorted(jobs or [], key=lambda item: str(item.get("job_ref") or ""))
    ]


def _row_token(payload, signing_key):
    canonical = json.dumps(payload, ensure_ascii=False, sort_keys=True,
                           separators=(",", ":")).encode("utf-8")
    encoded = base64.urlsafe_b64encode(canonical).decode("ascii").rstrip("=")
    signature = hmac.new(
        str(signing_key or "").encode("utf-8"),
        b"channel-row-v1\n" + canonical,
        hashlib.sha256,
    ).hexdigest()
    return encoded + "." + signature


def _decode_row_token(value, signing_key):
    try:
        encoded, supplied = str(value or "").split(".", 1)
        canonical = base64.urlsafe_b64decode(encoded + "=" * (-len(encoded) % 4))
        expected = hmac.new(
            str(signing_key or "").encode("utf-8"),
            b"channel-row-v1\n" + canonical,
            hashlib.sha256,
        ).hexdigest()
        if not hmac.compare_digest(supplied, expected):
            raise ValueError
        payload = json.loads(canonical.decode("utf-8"))
    except Exception as exc:
        raise ValueError("A system row token is missing or invalid.") from exc
    return payload


def _attach_workbook_metadata(
    data, generated_date, signing_key, jobs=(), *, row_versions=None,
    row_job_refs=None, row_sources=None, channels=None,
):
    catalog = _catalog_snapshot(jobs)
    channel_catalog = sorted({
        str(value or "").strip()
        for value in (channels or CHANNELS)
        if str(value or "").strip()
    })
    row_versions = {str(key): int(value or 0) for key, value in (row_versions or {}).items()}
    row_job_refs = {str(key): str(value or "") for key, value in (row_job_refs or {}).items()}
    row_sources = {
        str(key): {
            "channel": str((value or {}).get("channel") or ""),
            "source_detail": str((value or {}).get("source_detail") or ""),
        }
        for key, value in (row_sources or {}).items()
    }
    metadata = {
        "artifact_type": WORKBOOK_ARTIFACT_TYPE,
        "artifact_version": WORKBOOK_META_VERSION,
        "schema_version": pipeline_schema.SCHEMA_VERSION,
        "generated_date": str(generated_date or "")[:10],
        "artifact_id": str(uuid.uuid4()),
        "job_catalog": json.dumps(catalog, ensure_ascii=False, sort_keys=True,
                                  separators=(",", ":")),
        "channel_catalog": json.dumps(channel_catalog, ensure_ascii=False,
                                      separators=(",", ":")),
    }
    date.fromisoformat(metadata["generated_date"])
    metadata["signature"] = _workbook_signature(metadata, signing_key)
    wb = load_workbook(io.BytesIO(data))
    data_ws = (wb[pipeline_schema.PIPELINE_TABLE_NAME]
               if pipeline_schema.PIPELINE_TABLE_NAME in wb.sheetnames else wb[wb.sheetnames[0]])
    headers = {str(cell.value or ""): cell.column for cell in data_ws[1]}
    if not catalog and headers.get("Job"):
        seen = sorted({str(data_ws.cell(row=row, column=headers["Job"]).value or "").strip()
                       for row in range(2, data_ws.max_row + 1)
                       if str(data_ws.cell(row=row, column=headers["Job"]).value or "").strip()})
        catalog = [{"job_ref": "legacy-title-" + hashlib.sha256(title.encode("utf-8")).hexdigest()[:16],
                    "title": title, "catalog_revision": 1, "accept_new": True} for title in seen]
        metadata["job_catalog"] = json.dumps(catalog, ensure_ascii=False, sort_keys=True,
                                             separators=(",", ":"))
        metadata["signature"] = _workbook_signature(
            {key: value for key, value in metadata.items() if key != "signature"}, signing_key)
    token_col = headers.get("System Row Token")
    ref_col = headers.get("Row Ref")
    candidate_col = headers.get("System ID")
    if not token_col or not ref_col:
        raise ValueError("Workbook row identity columns are missing")
    for row_index in range(2, data_ws.max_row + 1):
        row_ref = str(data_ws.cell(row=row_index, column=ref_col).value or "")
        if not row_ref:
            continue
        payload = {
            "artifact_id": metadata["artifact_id"],
            "schema_version": pipeline_schema.SCHEMA_VERSION,
            "row_ref": row_ref,
            "candidate_id": str(data_ws.cell(row=row_index, column=candidate_col).value or "") if candidate_col else "",
            "record_version": row_versions.get(row_ref, 0),
            "job_ref": row_job_refs.get(row_ref, ""),
            "source_channel": (row_sources.get(row_ref) or {}).get("channel", ""),
            "source_detail": (row_sources.get(row_ref) or {}).get("source_detail", ""),
        }
        data_ws.cell(row=row_index, column=token_col, value=_row_token(payload, signing_key))
    ws = wb.create_sheet(WORKBOOK_META_SHEET)
    for row_index, key in enumerate(
        ("artifact_type", "artifact_version", "schema_version", "generated_date",
         "artifact_id", "job_catalog", "channel_catalog", "signature"), start=1
    ):
        ws.cell(row=row_index, column=1, value=key)
        ws.cell(row=row_index, column=2, value=metadata[key])
    ws.sheet_state = "veryHidden"
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _verify_workbook_metadata(data, expected_date, signing_key):
    _validate_xlsx_container(data)
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False, keep_links=False)
    if WORKBOOK_META_SHEET not in wb.sheetnames:
        raise ValueError(
            "This workbook has no Nexus provenance record. Download a fresh workbook today."
        )
    ws = wb[WORKBOOK_META_SHEET]
    metadata = {
        str(ws.cell(row=row, column=1).value or ""): str(
            ws.cell(row=row, column=2).value or ""
        )
        for row in range(1, 9)
    }
    signature = metadata.pop("signature", "")
    required = {
        "artifact_type", "artifact_version", "schema_version", "generated_date",
        "artifact_id", "job_catalog", "channel_catalog"
    }
    if set(metadata) != required:
        raise ValueError("The workbook provenance record is incomplete.")
    if metadata["artifact_type"] != WORKBOOK_ARTIFACT_TYPE:
        raise ValueError("This is not a Channel Analytics Candidate Pipeline workbook.")
    if metadata["artifact_version"] != WORKBOOK_META_VERSION:
        raise ValueError("This workbook format is no longer supported. Download a fresh workbook.")
    if metadata["schema_version"] != pipeline_schema.SCHEMA_VERSION:
        raise ValueError("This workbook schema is outdated. Download a fresh workbook.")
    try:
        uuid.UUID(metadata["artifact_id"])
        generated_date = date.fromisoformat(metadata["generated_date"])
        current_date = date.fromisoformat(str(expected_date or "")[:10])
    except (ValueError, TypeError) as exc:
        raise ValueError("The workbook provenance date or artifact ID is invalid.") from exc
    expected_signature = _workbook_signature(metadata, signing_key)
    if not hmac.compare_digest(signature, expected_signature):
        raise ValueError("The workbook provenance signature is invalid.")
    if generated_date != current_date:
        raise ValueError(
            "This workbook was generated on %s; the current Asia/Kolkata date is %s. "
            "Download a fresh workbook before submitting."
            % (generated_date.isoformat(), current_date.isoformat())
        )
    try:
        metadata["job_catalog"] = json.loads(metadata["job_catalog"])
        metadata["channel_catalog"] = json.loads(metadata["channel_catalog"])
        if not isinstance(metadata["job_catalog"], list) or not isinstance(
                metadata["channel_catalog"], list):
            raise ValueError
    except (TypeError, ValueError) as exc:
        raise ValueError("The workbook catalog snapshot is invalid.") from exc
    return metadata

# 候选人联系表用到的下拉
CANDIDATE_STATUS = ["未联系", "已联系", "无法联系", "不合适"]
CANDIDATE_SOURCES = ["RecruitEm", "领英公开", "GitHub"] + CHANNELS


# ---------------- 列定义 ----------------
def _col(key, header, kind="text", choices=None, aliases=None, minv=None, sys=False,
         hidden=False, lock_existing=False):
    """kind: text / int / date / choice。choices 给 choice 列的下拉项；minv 给 int 列下限。sys=系统列（表头灰+勿填批注）。"""
    return {"key": key, "header": header, "kind": kind,
            "choices": choices, "aliases": aliases or [], "minv": minv,
            "sys": sys, "hidden": hidden, "lock_existing": lock_existing}


def channel_columns(job_titles):
    columns = []
    for spec in pipeline_schema.MANUAL_COLUMNS:
        choices = CHANNELS if spec["key"] == "channel" else job_titles if spec["key"] == "job" else None
        columns.append(_col(
            spec["key"], spec["header"], spec["kind"], choices=choices,
            aliases=list(spec.get("aliases", ())), minv=0 if spec["kind"] == "int" else None,
        ))
    return columns


def candidate_columns(job_titles, sources=None):
    sources = sources or CANDIDATE_SOURCES
    return [
        _col("candidate_ref", "候选人ID", "text"),          # 系统预填的稳定引用（可空）
        _col("name", "候选人", "text"),                      # 系统预填
        _col("source", "来源渠道", "choice", choices=sources),
        _col("job", "关联职位", "choice", choices=job_titles),
        _col("region", "地区", "text"),
        _col("experience", "经验", "choice", choices=["有", "无", "未知"]),
        _col("intention", "求职意向", "choice", choices=["有", "无", "未知"]),
        _col("contact_status", "联系状态", "choice", choices=CANDIDATE_STATUS),  # HR 拉下拉标
        _col("contact_note", "联系方式/备注", "text"),
        _col("contact_date", "联系日期", "date"),
    ]


# ---------------- 生成（通用） ----------------
def build_xlsx(columns, prefill_rows=None, sheet_title="录入", extra_blank=0, blank_defaults=None):
    """按列定义生成 xlsx：表头 + choice 列下拉 + 预置行 + 若干带默认值的空行。返回字节。"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append([c["header"] for c in columns])
    head_fill = PatternFill("solid", fgColor="EEF1F5")
    sys_fill = PatternFill("solid", fgColor="D9DEE5")
    for i, c in enumerate(columns):
        cell = ws.cell(row=1, column=i + 1)
        cell.font = Font(bold=True, color=("8A94A0" if c.get("sys") else "000000"))
        cell.fill = sys_fill if c.get("sys") else head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if c.get("sys"):
            cell.comment = Comment(
                "System-owned field. HR cannot edit it; the service updates it when required.",
                "Nexus",
            )

    # 所有 choice 列的下拉项放隐藏引用表，避免内联列表 255 字符限制
    refs = wb.create_sheet("_refs")
    refs.sheet_state = "hidden"
    dv_map = {}
    ref_col = 1
    for ci, c in enumerate(columns):
        if c["kind"] == "choice" and c["choices"]:
            for i, opt in enumerate(c["choices"], 1):
                refs.cell(row=i, column=ref_col, value=opt)
            letter = get_column_letter(ref_col)
            # WPS and some Excel versions do not reliably render a dropdown
            # whose validation formula directly references a hidden sheet,
            # especially when the list currently contains only one Open job.
            # A workbook-level defined name is portable across Excel and WPS
            # and keeps the source list hidden from HR.
            list_name = "_nexus_choice_%d" % (ref_col,)
            wb.defined_names.add(DefinedName(
                list_name,
                attr_text="'_refs'!$%s$1:$%s$%d" % (
                    letter, letter, len(c["choices"]),
                ),
            ))
            dv = DataValidation(
                type="list", formula1="=" + list_name,
                allow_blank=True, showDropDown=False,
            )
            dv.promptTitle = c["header"]
            dv.prompt = "Select a value from the approved list."
            dv.showInputMessage = True
            ws.add_data_validation(dv)
            dv_map[ci] = dv
            ref_col += 1

    r = 2
    existing_rows = set()
    for row in (prefill_rows or []):
        if row.get("cand_id"):
            existing_rows.add(r)
        for ci, c in enumerate(columns):
            v = row.get(c["key"])
            if v is not None and v != "":
                cell = ws.cell(row=r, column=ci + 1)
                if c["kind"] == "date" and isinstance(v, str):
                    try:
                        v = date.fromisoformat(v[:10])
                    except ValueError:
                        pass
                cell.value = _safe_excel_value(v, c["kind"])
                if c["kind"] == "date":
                    cell.number_format = "yyyy-mm-dd"
        r += 1
    for _ in range(extra_blank):
        for ci, c in enumerate(columns):
            v = (blank_defaults or {}).get(c["key"])
            if v is not None:
                ws.cell(row=r, column=ci + 1, value=v)
        r += 1

    last = max(r - 1, 2)
    for ci, dv in dv_map.items():
        letter = get_column_letter(ci + 1)
        dv.add("%s2:%s%d" % (letter, letter, last))

    # Excel can enforce the non-Other half of the dependent source rule at
    # edit time. (The required-when-Other half remains fail-closed on submit.)
    key_to_column = {column["key"]: index + 1 for index, column in enumerate(columns)}
    if "channel" in key_to_column and "source_detail" in key_to_column:
        channel_letter = get_column_letter(key_to_column["channel"])
        detail_letter = get_column_letter(key_to_column["source_detail"])
        detail_validation = DataValidation(
            type="custom",
            formula1='=OR($%s2="Other",%s2="")' % (channel_letter, detail_letter),
            allow_blank=True,
        )
        detail_validation.error = (
            "Only fill Other Source Detail when Source Channel is Other."
        )
        detail_validation.errorTitle = "Source Channel mismatch"
        detail_validation.showErrorMessage = True
        ws.add_data_validation(detail_validation)
        detail_validation.add("%s2:%s%d" % (detail_letter, detail_letter, last))

    # System-owned fields are locked for HR but remain writable by the service.
    # Candidate identity is also locked for existing records while blank rows
    # remain available for HR to enter external candidates.
    for row_index in range(2, last + 1):
        for column_index, column in enumerate(columns, 1):
            ws.cell(row=row_index, column=column_index).protection = Protection(
                locked=bool(
                    column.get("sys")
                    or (column.get("lock_existing") and row_index in existing_rows)
                )
            )
    ws.protection.sheet = True

    preferred_widths = {
        "Candidate": 24,
        "Entry Date": 14,
        "Source Channel": 22,
        pipeline_schema.OTHER_SOURCE_DETAIL: 38,
        "Job": 24,
        "Current Stage": 24,
        pipeline_schema.STAGE_STARTED_ON: 20,
        "HR Owner": 18,
        "Rejection Reason": 28,
        "Note": 34,
    }
    for i, c in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = preferred_widths.get(
            c["header"], 20 if c["kind"] == "text" else 15
        )
        if c.get("hidden"):
            ws.column_dimensions[get_column_letter(i)].hidden = True
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(columns)), last)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------------- 解析（通用） ----------------
def _s(v):
    return ("" if v is None else str(v)).strip()


def _as_int(v):
    s = _s(v)
    if not s:
        return 0
    try:
        return int(float(s))
    except ValueError:
        return 0


def _table(data, filename):
    """xlsx/csv → 二维列表（含表头行）。"""
    if (filename or "").lower().endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        return [row for row in csv.reader(io.StringIO(text))]
    wb = load_workbook(io.BytesIO(data), data_only=True, keep_links=False)
    ws = None
    for name in wb.sheetnames:
        if name != "_refs" and wb[name].sheet_state == "visible":
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    return [[c.value for c in row] for row in ws.iter_rows()]


def parse_rows(columns, data, filename, required=None, skip=None):
    """通用解析：按列定义映射表头、按 kind 转换、choice/min 校验、跳过空/预置行。
    返回 {rows:[{key:值, __line__}], skipped:int, errors:[逐行问题]}。"""
    table = _table(data, filename)
    if not table:
        return {"rows": [], "skipped": 0, "errors": ["文件是空的"]}

    header = [_s(x) for x in table[0]]
    by_key = {}
    for i, h in enumerate(header):
        for c in columns:
            if h == c["header"] or h in c["aliases"]:
                by_key.setdefault(c["key"], i)
                break
    missing = [k for k in (required or []) if k not in by_key]
    if missing:
        miss_h = [next(c["header"] for c in columns if c["key"] == k) for k in missing]
        return {"rows": [], "skipped": 0,
                "errors": ["表头缺少列：%s（请用系统生成的模板）" % "、".join(miss_h)]}

    col_by_key = {c["key"]: c for c in columns}
    rows, errors, skipped = [], [], 0
    for ln, raw in enumerate(table[1:], start=2):
        if not any(_s(x) for x in raw):
            continue
        rec = {}
        for key, i in by_key.items():
            val = raw[i] if i < len(raw) else None
            rec[key] = _as_int(val) if col_by_key[key]["kind"] == "int" else _s(val)
        if skip and skip(rec):
            skipped += 1
            continue
        errs = []
        for key, c in col_by_key.items():
            if key not in rec:
                continue
            if c["kind"] == "choice" and rec[key] and c["choices"] and rec[key] not in c["choices"]:
                errs.append("第%d行：%s「%s」不在预设项" % (ln, c["header"], rec[key]))
            if c["kind"] == "int" and c["minv"] is not None and rec[key] < c["minv"]:
                errs.append("第%d行：%s 不能小于 %d" % (ln, c["header"], c["minv"]))
        if errs:
            errors += errs
            continue
        rec["__line__"] = ln
        rows.append(rec)
    return {"rows": rows, "skipped": skipped, "errors": errors}


# ---------------- 渠道汇总表（第二个程序，向后兼容） ----------------
def build_template_xlsx(jobs, day, by=""):
    cols = channel_columns([j["title"] for j in jobs])
    # Blank controlled rows are substantially easier than a jobs×channels zero
    # matrix. HR fills only combinations that actually had activity that day.
    return build_xlsx(cols, prefill_rows=[], sheet_title="未建档批量统计",
                      extra_blank=80, blank_defaults={"record_date": day, "filled_by": by})


def parse_sheet(data, filename, jobs, default_by="", default_date=None):
    title2id = {_s(j["title"]): j["id"] for j in jobs}
    cols = channel_columns([j["title"] for j in jobs])

    def skip(r):
        if not r.get("channel") and not r.get("job"):
            return True
        if (r.get("new_resumes", 0) == 0 and r.get("passed_screening", 0) == 0
                and r.get("recommended", 0) == 0 and r.get("rejected", 0) == 0
                and not r.get("note")):
            return True
        return False

    res = parse_rows(cols, data, filename, required=["channel", "job"], skip=skip)
    out, errors = [], list(res["errors"])
    for r in res["rows"]:
        jid = title2id.get(r.get("job"))
        if not jid:
            errors.append("第%d行：职位「%s」未找到" % (r.get("__line__", 0), r.get("job", "")))
            continue
        rd = (r.get("record_date") or (default_date or "")).strip()[:10]
        if not rd:
            errors.append("第%d行：缺日期" % r.get("__line__", 0))
            continue
        try:
            date.fromisoformat(rd)
        except ValueError:
            errors.append("第%d行：日期格式非法「%s」（应为 YYYY-MM-DD）" % (r.get("__line__", 0), rd))
            continue
        if r["channel"] == "Other" and not (r.get("source_detail") or "").strip():
            errors.append("第%d行：选择 Other 时必须填写其他来源说明" % r.get("__line__", 0))
            continue
        if r["channel"] != "Other" and (r.get("source_detail") or "").strip():
            errors.append(
                "Row %d: Other Source Detail is allowed only when Source Channel = Other"
                % r.get("__line__", 0)
            )
            continue
        out.append({"record_date": rd, "channel": r["channel"],
                    "source_detail": r.get("source_detail", ""), "job_request_id": jid,
                    "filled_by": r.get("filled_by") or default_by,
                    "new_resumes": r["new_resumes"], "passed_screening": r["passed_screening"],
                    "recommended": r["recommended"], "rejected": r["rejected"],
                    "note": r.get("note", "")})
    return {"rows": out, "skipped": res["skipped"], "errors": errors}


# ---------------- 候选人联系表（第一个程序，机制就绪；不落候选人库） ----------------
def build_candidate_template_xlsx(jobs, day, candidates=None, sources=None):
    """candidates: 系统爬到的候选人 [{candidate_ref,name,source,job,region,experience,intention}]（预填）。
    HR 只需在「联系状态」下拉里标已联系/无法联系等，并可补充联系方式/日期。"""
    cols = candidate_columns([j["title"] for j in jobs], sources)
    prefill = []
    for c in (candidates or []):
        prefill.append({k: c.get(k) for k in
                        ("candidate_ref", "name", "source", "job", "region", "experience", "intention")})
    return build_xlsx(cols, prefill_rows=prefill, sheet_title="候选人联系表",
                      extra_blank=(5 if candidates else 20), blank_defaults={})


def parse_candidate_sheet(data, filename, jobs, sources=None):
    """解析 HR 交回的候选人表 → 结构化行（含候选人身份 + HR 标的联系状态/备注/日期）。
    只产出行，不写库；上层核心 service 负责更新候选人状态并追加不可变 ContactActivity。"""
    cols = candidate_columns([j["title"] for j in jobs], sources)

    def skip(r):
        return not (r.get("candidate_ref") or r.get("name"))

    return parse_rows(cols, data, filename, required=["name"], skip=skip)


# ---------------- 招聘流水线候选人表（运营口径：每行一个候选人，状态走漏斗） ----------------
def pipeline_columns(job_titles, channels=None):
    channel_choices = list(channels or CHANNELS)
    columns = []
    for spec in pipeline_schema.columns_for("xlsx"):
        choices = (
            channel_choices if spec["key"] == "channel"
            else job_titles if spec["key"] == "job"
            else PIPELINE_STATUS if spec["key"] == "status"
            else None
        )
        columns.append(_col(
            spec["key"], spec["header"], spec["kind"], choices=choices,
            aliases=list(spec.get("aliases", ())), sys=bool(spec.get("system")),
            hidden=bool(spec.get("hidden")), lock_existing=bool(spec.get("lock_existing")),
        ))
    return columns


def build_pipeline_template_xlsx(jobs, day, by="", candidates=None, signing_key=""):
    """候选人跟进表：把在跟进中的候选人（带「记录ID」+当前状态）预填进去，HR 直接在「状态」列往前改；
    末尾留空行给新候选人（记录ID 留空）。上传时——有记录ID 的按 ID 原地更新，没 ID 的当新增，系统自动分辨。"""
    candidates = list(candidates or [])
    channel_choices = sorted({
        *CHANNELS,
        *(str(candidate.get("channel") or "").strip() for candidate in candidates),
    } - {""})
    cols = pipeline_columns(sorted({j["title"] for j in jobs}), channel_choices)
    id2title = {j["id"]: j["title"] for j in jobs}
    prefill = []
    row_versions, row_job_refs, row_sources = {}, {}, {}
    for c in candidates:
        row_ref = c.get("application_ref") or c.get("ext_ref") or ("candidate-%s" % c.get("id"))
        prefill.append({
            "cand_id": str(c.get("application_ref") or c.get("id") or ""),
            "name": c.get("name") or "",
            "channel": c.get("channel") or "",
            "source_detail": c.get("source_detail") or "",
            "job": c.get("job_title") or id2title.get(c.get("job_request_id"), ""),
            "status": c.get("status") or "New Lead",
            "rejection_reason": c.get("rejection_reason") or "",
            "note": c.get("note") or "",
            "filled_by": c.get("filled_by") or "",
            "row_ref": row_ref,
        })
        row_versions[str(row_ref)] = int(c.get("record_version") or 1)
        row_job_refs[str(row_ref)] = str(
            c.get("job_ref") or next(
                (job.get("job_ref") for job in jobs
                 if job.get("id") == c.get("job_request_id")),
                "",
            ) or ""
        )
        row_sources[str(row_ref)] = {
            "channel": c.get("channel") or "",
            "source_detail": c.get("source_detail") or "",
        }
    blank_count = 30 if prefill else 60
    # HR surfaces contain commands only. Entry Date is assigned when a row is
    # first imported, while Stage Started On comes from immutable stage events.
    prefill.extend({"row_ref": "manual-" + str(uuid.uuid4()), "filled_by": by}
                   for _ in range(blank_count))
    data = build_xlsx(cols, prefill_rows=prefill,
                      sheet_title=pipeline_schema.PIPELINE_TABLE_NAME,
                      extra_blank=0, blank_defaults={})
    return _attach_workbook_metadata(
        data, day, signing_key, jobs,
        row_versions=row_versions, row_job_refs=row_job_refs,
        row_sources=row_sources, channels=channel_choices,
    )


def parse_pipeline_sheet(data, filename, jobs, default_by="", default_date=None,
                         signing_key=""):
    """解析 HR 交回的候选人跟进表 → 候选人行。带「记录ID」= 已有候选人（更新），无 ID = 新候选人（新增）。"""
    metadata = _verify_workbook_metadata(data, default_date, signing_key)
    current_by_ref = {
        str(j.get("job_ref") or ("legacy-id-%s" % j.get("id"))): j for j in jobs
    }
    current_by_title = {_s(j.get("title")): j for j in jobs}
    catalog_by_ref = {
        str(j.get("job_ref") or ""): j for j in metadata["job_catalog"]
    }
    new_catalog_by_title = {}
    for job in metadata["job_catalog"]:
        if job.get("accept_new"):
            new_catalog_by_title.setdefault(_s(job.get("title")), []).append(job)
    cols = pipeline_columns(
        sorted({_s(j.get("title")) for j in metadata["job_catalog"] if _s(j.get("title"))}),
        metadata["channel_catalog"],
    )

    def skip(r):
        return not (r.get("channel") or r.get("name"))  # 渠道和姓名都空 -> 空行

    res = parse_rows(cols, data, filename, required=["channel"], skip=skip)
    out, errors = [], list(res["errors"])
    seen_refs, seen_tokens = {}, {}
    duplicate_identity = False
    for r in res["rows"]:
        try:
            token = _decode_row_token(r.get("row_token"), signing_key)
        except ValueError as exc:
            errors.append("Row %d: %s" % (r.get("__line__", 0), exc))
            continue
        if (token.get("artifact_id") != metadata["artifact_id"] or
                token.get("row_ref") != (r.get("row_ref") or "") or
                str(token.get("candidate_id") or "") != str(r.get("cand_id") or "")):
            errors.append("Row %d: signed system identity does not match the row" % r.get("__line__", 0))
            continue
        row_ref = (r.get("row_ref") or "").strip()
        row_token = (r.get("row_token") or "").strip()
        for value, seen, label in (
            (row_ref, seen_refs, "Row Ref"),
            (row_token, seen_tokens, "System Row Token"),
        ):
            if value in seen:
                errors.append(
                    "Rows %d and %d contain the same %s; copied system identities are not accepted"
                    % (seen[value], r.get("__line__", 0), label)
                )
                duplicate_identity = True
            else:
                seen[value] = r.get("__line__", 0)

        existing_row = bool((r.get("cand_id") or "").strip())
        if existing_row:
            catalog_ref = str(token.get("job_ref") or "")
            catalog_job = catalog_by_ref.get(catalog_ref)
            if not catalog_job:
                errors.append("Row %d: the signed Job Ref is missing from this artifact" % r.get("__line__", 0))
                continue
            if _s(r.get("job")) != _s(catalog_job.get("title")):
                errors.append("Row %d: Job is protected for an existing application" % r.get("__line__", 0))
                continue
            if (_s(r.get("channel")) != _s(token.get("source_channel")) or
                    _s(r.get("source_detail")) != _s(token.get("source_detail"))):
                errors.append("Row %d: Source attribution is protected for an existing application" % r.get("__line__", 0))
                continue
        else:
            choices = new_catalog_by_title.get(_s(r.get("job"))) or []
            if len(choices) != 1:
                errors.append(
                    "Row %d: Job must identify exactly one Open requisition in this signed workbook"
                    % r.get("__line__", 0)
                )
                continue
            catalog_job = choices[0]
            catalog_ref = str(catalog_job.get("job_ref") or "")

        current_job = current_by_ref.get(catalog_ref)
        if current_job is None and catalog_ref.startswith("legacy-title-"):
            current_job = current_by_title.get(_s(catalog_job.get("title")))
        if not current_job:
            errors.append("Row %d: Job no longer exists in the operational catalog" % r.get("__line__", 0))
            continue
        jid = current_job.get("id")
        # New applications can only enter an Open requisition. Existing rows
        # may finish work after a requisition moves to Paused/Closing/Closed.
        if not existing_row and _s(current_job.get("status") or "open").casefold() != "open":
            errors.append("Row %d: new candidates require an Open requisition" % r.get("__line__", 0))
            continue
        # Entry Date is system-owned. A workbook value is display-only and is
        # never trusted for a new or existing candidate.
        rd = (default_date or "").strip()[:10]
        if not rd:
            errors.append("第%d行：缺日期" % r.get("__line__", 0))
            continue
        try:
            date.fromisoformat(rd)
        except ValueError:
            errors.append("第%d行：日期格式非法「%s」（应为 YYYY-MM-DD）" % (r.get("__line__", 0), rd))
            continue
        if r["channel"] == "Other" and not (r.get("source_detail") or "").strip():
            errors.append("第%d行：选择 Other 时必须填写其他来源说明" % r.get("__line__", 0))
            continue
        if r["channel"] != "Other" and (r.get("source_detail") or "").strip():
            errors.append(
                "Row %d: Other Source Detail is allowed only when Source Channel = Other"
                % r.get("__line__", 0)
            )
            continue
        if (r.get("status") or "New Lead") == "Rejected" and not (r.get("rejection_reason") or "").strip():
            errors.append("第%d行：Rejected 必须填写 Rejection Reason" % r.get("__line__", 0))
            continue
        out.append({"cand_id": (r.get("cand_id") or "").strip(),
                    "row_ref": row_ref, "record_date": rd,
                    "name": r.get("name", ""), "channel": r["channel"],
                    "source_detail": r.get("source_detail", ""),
                    "job_request_id": jid,
                    "job_ref": (catalog_job or {}).get("job_ref", ""),
                    "catalog_revision": int((catalog_job or {}).get("catalog_revision") or 1),
                    "expected_version": int(token.get("record_version") or 0),
                    "artifact_id": metadata["artifact_id"],
                    "status": r.get("status") or "New Lead",
                    "stage_date": "", "rejection_reason": r.get("rejection_reason", ""),
                    "note": r.get("note", ""), "filled_by": r.get("filled_by") or default_by})
    if duplicate_identity:
        return {"rows": [], "skipped": res["skipped"], "errors": errors, "fatal": True,
                "artifact_id": metadata["artifact_id"]}
    return {"rows": out, "skipped": res["skipped"], "errors": errors,
            "artifact_id": metadata["artifact_id"]}


# ================= 招聘分析导出（汇报用 xlsx；供 /api/channel/export.xlsx） =================
def build_analytics_xlsx(a):
    """把 channel_report.analytics() 结果导出成多 sheet 工作簿。返回 bytes。"""
    wb = Workbook()
    hf = Font(bold=True)
    hfill = PatternFill("solid", fgColor="EEF1F5")
    gname = {"day": "日", "week": "周", "month": "月", "year": "年"}

    def pct(x):
        return "" if x is None else ("%.0f%%" % round(x * 100))

    sc = a["summary"]
    pc = a.get("prev_summary")

    def chg(key, rate=False):
        if not pc:
            return ""
        cur, prev = sc.get(key), pc.get(key)
        if rate:
            if cur is None or prev is None:
                return ""
            return "%+.1fpp" % ((cur - prev) * 100)
        if not prev:
            return ""
        return "%+.0f%%" % ((cur - prev) / prev * 100)

    # ---- 概览 ----
    ws = wb.active
    ws.title = "概览"
    ws.append(["招聘分析概览"]); ws["A1"].font = Font(bold=True, size=14)
    win = a["window"]
    ws.append(["窗口", "%s ~ %s" % (win["from"], win["to"])])
    ws.append(["粒度", gname.get(a["granularity"], a["granularity"])])
    pw = a.get("prev_window")
    if pw:
        ws.append(["上一周期", "%s ~ %s" % (pw["from"], pw["to"])])
    ws.append(["口径", "人工录入（未逐人建档）· 时区 Asia/Kolkata"])
    ws.append([])
    ws.append(["指标", "本期", "上一周期", "环比"])
    for c in ws[ws.max_row]:
        c.font = hf; c.fill = hfill
    disp = [
        ("新增简历", sc["new"], (pc["new"] if pc else ""), chg("new")),
        ("初筛通过", sc["passed"], (pc["passed"] if pc else ""), chg("passed")),
        ("推荐", sc["recommended"], (pc["recommended"] if pc else ""), chg("recommended")),
        ("拒绝", sc["rejected"], (pc["rejected"] if pc else ""), chg("rejected")),
        ("初筛通过率", pct(sc["conversion"]), (pct(pc["conversion"]) if pc else ""), chg("conversion", True)),
        ("推荐率", pct(sc["recommend_rate"]), (pct(pc["recommend_rate"]) if pc else ""), chg("recommend_rate", True)),
        ("简历/天", sc["resumes_per_day"], (pc["resumes_per_day"] if pc else ""), chg("resumes_per_day")),
        ("推荐/周", sc["recommended_per_week"], (pc["recommended_per_week"] if pc else ""), chg("recommended_per_week")),
        ("简历量进度", pct(sc["resume_target_progress"]), "", ""),
    ]
    for r in disp:
        ws.append(list(r))
    ws.append([])
    ws.append(["自动摘要"]); ws[ws.max_row][0].font = hf
    for line in a.get("insights", []):
        ws.append([line])

    # ---- 渠道明细 ----
    ws2 = wb.create_sheet("渠道明细")
    hascost = a.get("has_cost")
    head2 = ["渠道", "新增", "占比", "初筛", "转化率", "推荐", "推荐率", "拒绝"]
    if hascost:
        head2 += ["成本", "每份成本"]
    ws2.append(head2)
    for c in ws2[1]:
        c.font = hf; c.fill = hfill
    for c in a["channels"]:
        row = [c["channel"], c["new"], pct(c["share"]), c["passed"], pct(c["conversion"]),
               c["recommended"], pct(c["recommend_rate"]), c["rejected"]]
        if hascost:
            row += [c.get("cost") or 0, c.get("cost_per_resume") if c.get("cost_per_resume") is not None else ""]
        ws2.append(row)

    # ---- 趋势 ----
    ws3 = wb.create_sheet("趋势")
    ws3.append(["时间", "新增", "初筛", "推荐", "拒绝", "转化率", "推荐率"])
    for c in ws3[1]:
        c.font = hf; c.fill = hfill
    for b in a["timeseries"]:
        ws3.append([b["label"], b["new"], b["passed"], b["recommended"], b["rejected"],
                    pct(b["conversion"]), pct(b["recommend_rate"])])

    # ---- 职位进度 ----
    ws4 = wb.create_sheet("职位进度")
    ws4.append(["职位", "窗口新增", "目标简历量", "简历量进度", "窗口推荐", "目标录用"])
    for c in ws4[1]:
        c.font = hf; c.fill = hfill
    for j in a["jobs"]:
        ws4.append([j["title"], j["window_new"], j["target_resumes"], pct(j["resume_progress"]),
                    j["window_recommended"], j["target_headcount"]])

    for wsx in (ws, ws2, ws3, ws4):
        for col in wsx.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            wsx.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 42)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
