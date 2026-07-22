import io
import sqlite3
import zipfile
from pathlib import Path

from openpyxl import load_workbook

import channel_pipeline_schema as schema
import db as nexus_db
import sheet_io


KEY = "v14-job-catalog-row-signing-key-2026-07-22"


def _filled(workbook, *, job_title="CSR"):
    wb = load_workbook(io.BytesIO(workbook))
    ws = wb[schema.PIPELINE_TABLE_NAME]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["Candidate"], "Candidate One")
    ws.cell(2, headers["Source Channel"], "LinkedIn")
    ws.cell(2, headers["Job"], job_title)
    ws.cell(2, headers["Current Stage"], "New Lead")
    out = io.BytesIO(); wb.save(out); return out.getvalue()


def main():
    original = [{"id": 1, "job_ref": "REQ-CSR", "title": "CSR",
                 "status": "open", "catalog_revision": 3}]
    workbook = sheet_io.build_pipeline_template_xlsx(
        original, "2026-07-22", "HR-01", [], signing_key=KEY)

    # Every blank HR input row must retain a visible Job dropdown in both
    # Microsoft Excel and WPS, including the common single-Open-job case.
    dropdown_wb = load_workbook(io.BytesIO(workbook))
    dropdown_ws = dropdown_wb[schema.PIPELINE_TABLE_NAME]
    dropdown_headers = {cell.value: cell.column for cell in dropdown_ws[1]}
    job_cell = dropdown_ws.cell(3, dropdown_headers["Job"])
    job_validations = [
        validation for validation in dropdown_ws.data_validations.dataValidation
        if job_cell.coordinate in validation.cells
    ]
    assert len(job_validations) == 1
    assert str(job_validations[0].formula1).startswith("=_nexus_choice_")
    assert job_validations[0].showDropDown is False
    assert job_cell.protection.locked is False
    assert str(job_validations[0].formula1)[1:] in dropdown_wb.defined_names

    # A rename after download is safe because the signed catalog resolves by
    # stable job_ref, never by the current display title.
    renamed = [{"id": 1, "job_ref": "REQ-CSR", "title": "Customer Service Representative",
                "status": "open", "catalog_revision": 4}]
    parsed = sheet_io.parse_pipeline_sheet(
        _filled(workbook), "ChannelAnalytics_20260722.xlsx", renamed,
        "HR-01", "2026-07-22", signing_key=KEY)
    assert not parsed["errors"] and parsed["rows"][0]["job_request_id"] == 1
    assert parsed["rows"][0]["job_ref"] == "REQ-CSR"

    # A job closed after workbook generation rejects new rows but does not
    # invalidate the artifact or destroy prior history.
    closed = [{**renamed[0], "status": "closed"}]
    rejected = sheet_io.parse_pipeline_sheet(
        _filled(workbook), "ChannelAnalytics_20260722.xlsx", closed,
        "HR-01", "2026-07-22", signing_key=KEY)
    assert any("Open requisition" in error for error in rejected["errors"])

    # HR commands are editable, but signed system identity is not.
    tampered = load_workbook(io.BytesIO(_filled(workbook)))
    ws = tampered[schema.PIPELINE_TABLE_NAME]
    headers = {cell.value: cell.column for cell in ws[1]}
    ws.cell(2, headers["Row Ref"], "forged-row")
    out = io.BytesIO(); tampered.save(out)
    bad = sheet_io.parse_pipeline_sheet(
        out.getvalue(), "ChannelAnalytics_20260722.xlsx", original,
        "HR-01", "2026-07-22", signing_key=KEY)
    assert any("signed system identity" in error for error in bad["errors"])

    # Existing applications keep their frozen requisition/source identity and
    # remain actionable after the requisition closes.  Their true optimistic
    # concurrency version is carried in the signed row token.
    historical_workbook = sheet_io.build_pipeline_template_xlsx(
        closed, "2026-07-22", "HR-01", [{
            "id": 99, "application_ref": "APP-HISTORICAL-99",
            "name": "Historical Candidate", "channel": "LinkedIn",
            "source_detail": "", "job_request_id": 1,
            "job_ref": "REQ-CSR", "job_title": renamed[0]["title"],
            "status": "HR Screening", "record_version": 7,
        }], signing_key=KEY,
    )
    historical = sheet_io.parse_pipeline_sheet(
        historical_workbook, "ChannelAnalytics_20260722.xlsx", closed,
        "HR-01", "2026-07-22", signing_key=KEY,
    )
    assert not historical["errors"] and len(historical["rows"]) == 1
    assert historical["rows"][0]["expected_version"] == 7
    assert historical["rows"][0]["job_ref"] == "REQ-CSR"

    # First-touch attribution is immutable. Editing a visible source field
    # cannot change the authority carried by the signed system row token.
    source_tamper = load_workbook(io.BytesIO(historical_workbook))
    source_ws = source_tamper[schema.PIPELINE_TABLE_NAME]
    source_headers = {cell.value: cell.column for cell in source_ws[1]}
    source_ws.cell(2, source_headers["Source Channel"], "Naukri")
    out = io.BytesIO(); source_tamper.save(out)
    source_bad = sheet_io.parse_pipeline_sheet(
        out.getvalue(), "ChannelAnalytics_20260722.xlsx", closed,
        "HR-01", "2026-07-22", signing_key=KEY,
    )
    assert any("Source attribution is protected" in error for error in source_bad["errors"])

    # Copying an existing row duplicates its authority. Reject the entire
    # artifact instead of applying one ambiguous identity twice.
    duplicate = load_workbook(io.BytesIO(historical_workbook))
    duplicate_ws = duplicate[schema.PIPELINE_TABLE_NAME]
    for column in range(1, duplicate_ws.max_column + 1):
        duplicate_ws.cell(3, column, duplicate_ws.cell(2, column).value)
    out = io.BytesIO(); duplicate.save(out)
    duplicate_result = sheet_io.parse_pipeline_sheet(
        out.getvalue(), "ChannelAnalytics_20260722.xlsx", closed,
        "HR-01", "2026-07-22", signing_key=KEY,
    )
    assert duplicate_result.get("fatal") is True
    assert not duplicate_result["rows"]
    assert any("copied system identities" in error for error in duplicate_result["errors"])

    # HR uploads accept only ordinary XLSX packages. Embedded external links
    # and macro payloads are rejected before openpyxl parses cell content.
    malicious = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(workbook), "r") as source, zipfile.ZipFile(
            malicious, "w", zipfile.ZIP_DEFLATED) as target:
        for member in source.infolist():
            target.writestr(member, source.read(member.filename))
        target.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")
    rejected_container = False
    try:
        sheet_io.parse_pipeline_sheet(
            malicious.getvalue(), "ChannelAnalytics_20260722.xlsx", original,
            "HR-01", "2026-07-22", signing_key=KEY,
        )
    except ValueError as exc:
        rejected_container = "External links and macros" in str(exc)
    assert rejected_container

    sql = Path("schema.sql").read_text(encoding="utf-8")
    assert "candidate_application" in sql
    assert "record_type='search_profile'" in sql
    assert "ON DELETE RESTRICT" in sql
    assert "ON CONFLICT (candidate_id, job_request_id) DO NOTHING" not in sql
    assert "APP-RECOVER-" in sql
    assert "schema_migration_anomaly" in sql
    assert "a.job_request_id IS NOT DISTINCT FROM c.job_request_id" in sql
    assert "JOIN LATERAL" in sql
    assert "(ca.job_request_id IS NOT DISTINCT FROM c.job_request_id) DESC" in sql
    assert "ON CONFLICT DO NOTHING" in sql
    assert "APP-LEGACY-" in sql
    for transaction_unsafe_sql in (
        "CONCURRENTLY", "VACUUM", "REINDEX", "CREATE DATABASE",
        "DROP DATABASE", "ALTER SYSTEM",
    ):
        assert transaction_unsafe_sql not in sql.upper()
    db_source = Path("db.py").read_text(encoding="utf-8")
    assert "SELECT pg_advisory_xact_lock(%s)" in db_source
    assert "job_request_id IS NOT DISTINCT FROM %s" in db_source
    nexus_db._validate_application_stage_change("New Lead", "Interview 1")
    nexus_db._validate_application_stage_change("On Hold", "HR Screening")
    for current_stage, next_stage in (
        ("Interview 2 / Final", "HR Screening"),
        ("Rejected", "New Lead"),
        ("Hired", "On Hold"),
    ):
        blocked = False
        try:
            nexus_db._validate_application_stage_change(current_stage, next_stage)
        except ValueError:
            blocked = True
        assert blocked

    # Reproduce the production failure mode with SQL NULL uniqueness rules:
    # a partial migration already contains APP-0000000001 for candidate 1 and
    # a restart attempts the same legacy row. Targeting only the nullable
    # composite key raises on application_ref; targetless DO NOTHING is safe.
    migration_db = sqlite3.connect(":memory:")
    migration_db.execute("""CREATE TABLE app(
        application_ref TEXT NOT NULL UNIQUE,
        candidate_id INTEGER NOT NULL,
        job_request_id INTEGER,
        UNIQUE(candidate_id, job_request_id))""")
    migration_db.execute("INSERT INTO app VALUES('APP-0000000001',1,NULL)")
    targeted_failed = False
    try:
        migration_db.execute("""INSERT INTO app VALUES('APP-0000000001',1,NULL)
            ON CONFLICT(candidate_id,job_request_id) DO NOTHING""")
    except sqlite3.IntegrityError:
        targeted_failed = True
    assert targeted_failed
    migration_db.execute("""INSERT INTO app VALUES('APP-0000000001',1,NULL)
        ON CONFLICT DO NOTHING""")
    assert migration_db.execute("SELECT COUNT(*) FROM app").fetchone()[0] == 1

    # An old application can legitimately remain on the previous requisition
    # after the legacy candidate projection moves. Add the missing current
    # pair without mutating the old workflow, and make the backfill restartable.
    migration_db.execute("""INSERT INTO app VALUES(
        'APP-LEGACY-0000000001-7',1,7) ON CONFLICT DO NOTHING""")
    migration_db.execute("""INSERT INTO app VALUES(
        'APP-LEGACY-0000000001-7',1,7) ON CONFLICT DO NOTHING""")
    assert migration_db.execute(
        "SELECT COUNT(*) FROM app WHERE candidate_id=1 AND job_request_id=7"
    ).fetchone()[0] == 1
    assert migration_db.execute(
        "SELECT COUNT(*) FROM app WHERE candidate_id=1"
    ).fetchone()[0] == 2
    migration_db.close()

    # Schema initialization is serialized and atomic. The success path commits
    # once; any schema error rolls back and still closes the connection.
    class FakeCursor:
        def __init__(self, connection):
            self.connection = connection

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def execute(self, statement, params=None):
            self.connection.calls.append((statement, params))
            if self.connection.fail_schema and len(self.connection.calls) == 4:
                raise RuntimeError("simulated schema failure")

        def fetchone(self):
            return self.connection.applied_row

    class FakeConnection:
        def __init__(self, fail_schema=False, applied_row=None):
            self.autocommit = True
            self.fail_schema = fail_schema
            self.calls = []
            self.applied_row = applied_row
            self.commits = self.rollbacks = self.closes = 0

        def cursor(self):
            return FakeCursor(self)

        def commit(self):
            self.commits += 1

        def rollback(self):
            self.rollbacks += 1

        def close(self):
            self.closes += 1

    original_get_conn = nexus_db.get_conn
    try:
        successful_connection = FakeConnection()
        nexus_db.get_conn = lambda: successful_connection
        nexus_db.init_db()
        assert successful_connection.autocommit is False
        assert "pg_advisory_xact_lock" in successful_connection.calls[0][0]
        assert "schema_migrations" in successful_connection.calls[1][0]
        assert successful_connection.calls[-1][0].startswith("INSERT INTO schema_migrations")
        assert successful_connection.commits == 1
        assert successful_connection.rollbacks == 0
        assert successful_connection.closes == 1

        failed_connection = FakeConnection(fail_schema=True)
        nexus_db.get_conn = lambda: failed_connection
        failed = False
        try:
            nexus_db.init_db()
        except RuntimeError as exc:
            failed = "simulated schema failure" in str(exc)
        assert failed
        assert failed_connection.commits == 0
        assert failed_connection.rollbacks == 1
        assert failed_connection.closes == 1

        schema_checksum = __import__("hashlib").sha256(sql.encode("utf-8")).hexdigest()
        already_applied = FakeConnection(applied_row=(schema_checksum,))
        nexus_db.get_conn = lambda: already_applied
        nexus_db.init_db()
        assert len(already_applied.calls) == 3
        assert already_applied.commits == 1

        checksum_conflict = FakeConnection(applied_row=("0" * 64,))
        nexus_db.get_conn = lambda: checksum_conflict
        conflict_raised = False
        try:
            nexus_db.init_db()
        except RuntimeError as exc:
            conflict_raised = "immutable migration version" in str(exc)
        assert conflict_raised
        assert checksum_conflict.rollbacks == 1
    finally:
        nexus_db.get_conn = original_get_conn
    html = Path("panel.html").read_text(encoding="utf-8")
    assert "Operational Job Requisitions" in html
    assert "Advanced: link a Talent Discovery search profile" in html
    assert "SEARCH_PROFILES.filter(p=>p.status==='open')" in html
    assert "正式配额" not in html[html.index('id="view-jobs"'):html.index('id="view-docs"')]
    print("Operational/Search Profile separation and application migration: PASSED")
    print("Frozen job catalog, rename compatibility, closure guard, row HMAC: PASSED")


if __name__ == "__main__":
    main()
