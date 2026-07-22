import io
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

import bot
import channel_pipeline_schema as pipeline_schema
import sheet_io


class FakeChannelStore:
    def __init__(self):
        self.rows = {}
        self.settings = {
            "lark_channel_app_token": "app-token",
            "lark_channel_pipeline_table_id": "pipeline-table",
            "lark_channel_manual_table_id": "manual-table",
            "lark_channel_url": "https://example.test/base/channel?table=default-table&view=default-view",
            "lark_channel_schema_version": "channel-analytics-v2",
        }
        self.candidates = {}
        self.next_candidate_id = 1

    def upsert(self, record_date, channel, job_request_id, filled_by,
               new_resumes, passed_screening, recommended, rejected, note,
               source_detail=""):
        self.rows[(record_date, channel, source_detail, job_request_id)] = {
            "new_resumes": new_resumes, "passed_screening": passed_screening,
            "recommended": recommended, "rejected": rejected, "filled_by": filled_by,
        }

    def create_candidate(self, apply_date, name, channel, job_request_id, status,
                         note, filled_by, source, ext_ref, lark_record_id, source_detail):
        candidate_id = self.next_candidate_id
        self.next_candidate_id += 1
        self.candidates[candidate_id] = {"id": candidate_id, "apply_date": apply_date,
            "name": name, "channel": channel, "source_detail": source_detail,
            "job_request_id": job_request_id, "status": status, "note": note,
            "filled_by": filled_by, "ext_ref": ext_ref, "lark_record_id": lark_record_id}
        return candidate_id

    def transition(self, candidate_id, to_stage, effective_date, changed_by,
                   rejection_reason, note, event_ref):
        old = self.candidates[candidate_id]["status"]
        self.candidates[candidate_id]["status"] = to_stage
        self.candidates[candidate_id]["stage_date"] = effective_date
        return {"from_stage": old, "to_stage": to_stage, "event_ref": event_ref or "generated"}


def main():
    workbook_key = "channel-route-workbook-signing-key-2026"
    store = FakeChannelStore()
    jobs = [{"id": 7, "title": "Sales", "target_headcount": 1, "target_resume_count": 10}]
    bot.PANEL_PASSWORD = "channel-route-password"
    bot.PUBLIC_BASE_URL = "https://nexus.example.test"
    bot.NEXUS_INTEGRATION_SIGNING_KEY = workbook_key
    bot._kolkata_today = lambda: date(2026, 7, 21)
    bot.db.list_job_requests = lambda only_open=False: jobs
    bot.db.list_candidates_active = lambda: []
    bot.db.list_candidate_applications_active = lambda: []
    bot.db.get_settings = lambda: store.settings
    bot.db.set_setting = lambda key, value: store.settings.__setitem__(key, value)
    bot.db.upsert_channel_record = store.upsert
    bot.db.is_admin = lambda open_id: open_id == "admin-open-id"
    bot.db.get_job_request = lambda job_id: jobs[0] if job_id == 7 else None
    bot.db.create_candidate = store.create_candidate
    bot.db.get_candidate = lambda candidate_id: store.candidates.get(candidate_id)
    bot.db.get_candidate_by_ext_ref = lambda ref: next((c for c in store.candidates.values() if c.get("ext_ref") == ref), None)
    bot.db.update_candidate = lambda candidate_id, **fields: store.candidates[candidate_id].update(fields)
    bot.db.transition_candidate_stage = store.transition
    def create_application(entry_date, name, channel, job_request_id, note="", hr_owner="",
                           source="manual", external_ref="", lark_record_id="", source_detail=""):
        existing = next((row for row in store.candidates.values()
                         if external_ref and row.get("ext_ref") == external_ref
                         and row.get("job_request_id") == job_request_id), None)
        if existing:
            candidate_id = existing["id"]
            return ({"id": candidate_id, "application_ref": "APP-%d" % candidate_id,
                     "candidate_id": candidate_id, "current_stage": existing["status"]}, False)
        candidate_id = store.create_candidate(entry_date, name, channel, job_request_id,
            "New Lead", note, hr_owner, source, external_ref, lark_record_id, source_detail)
        return ({"id": candidate_id, "application_ref": "APP-%d" % candidate_id,
                 "candidate_id": candidate_id, "current_stage": "New Lead"}, True)
    bot.db.create_candidate_application = create_application
    bot.db.get_candidate_application = lambda app_id: ({
        **store.candidates[app_id], "application_ref": "APP-%d" % app_id,
        "candidate_id": app_id, "current_stage": store.candidates[app_id]["status"],
    } if app_id in store.candidates else None)
    bot.db.get_candidate_application_by_ref = lambda ref: (bot.db.get_candidate_application(int(ref.split('-')[1])) if ref.startswith('APP-') else None)
    bot.db.get_candidate_application_by_lark = lambda record_id: next(({
        **row, "application_ref": "APP-%d" % row["id"], "candidate_id": row["id"],
        "current_stage": row["status"]} for row in store.candidates.values()
        if row.get("lark_record_id") == record_id), None)
    bot.db.update_candidate_application = lambda ref, **fields: store.candidates[int(ref.split('-')[1])].update({
        ("status" if k == "current_stage" else "filled_by" if k == "hr_owner" else k): v
        for k, v in fields.items()
    })
    bot.db.transition_candidate_application = lambda ref, *args: store.transition(int(ref.split('-')[1]), *args)
    bot.db.list_candidate_application_stage_events = lambda app_id: []

    workbook = sheet_io.build_xlsx(
        sheet_io.pipeline_columns(["Sales"]),
        prefill_rows=[{
            "record_date": "2026-07-21", "name": "Asha", "channel": "LinkedIn",
            "source_detail": "", "job": "Sales", "status": "HR Screening",
            "stage_date": "2026-07-21", "rejection_reason": "",
            "note": "route", "filled_by": "ignored", "cand_id": "", "row_ref": "manual-test-row",
        }],
        sheet_title="未建档批量统计",
    )
    workbook = sheet_io._attach_workbook_metadata(
        workbook, "2026-07-21", workbook_key
    )
    check_wb = load_workbook(io.BytesIO(workbook))
    check_ws = check_wb["未建档批量统计"]
    headers = [cell.value for cell in check_ws[1]]
    assert check_ws.protection.sheet is True
    assert "Entry Date" not in headers
    assert "Stage Started On" not in headers
    assert check_ws.cell(2, headers.index("Current Stage") + 1).protection.locked is False
    assert headers.index(pipeline_schema.OTHER_SOURCE_DETAIL) == headers.index("Source Channel") + 1
    assert headers[:8] == [
        "Candidate", "Source Channel", pipeline_schema.OTHER_SOURCE_DETAIL,
        "Job", "Current Stage", "HR Owner", "Rejection Reason", "Note",
    ]
    assert check_ws.column_dimensions["I"].hidden is True
    assert check_ws.column_dimensions["J"].hidden is True
    detail_letter = check_ws.cell(1, headers.index(pipeline_schema.OTHER_SOURCE_DETAIL) + 1).column_letter
    assert any(detail_letter in str(validation.sqref)
               and '="Other"' in str(validation.formula1)
               for validation in check_ws.data_validations.dataValidation)

    blank_template = sheet_io.build_pipeline_template_xlsx(
        jobs, "2026-07-21", "HR-01", [], signing_key=workbook_key
    )
    blank_wb = load_workbook(io.BytesIO(blank_template), data_only=True)
    blank_ws = blank_wb["Candidate Pipeline"]
    blank_headers = [cell.value for cell in blank_ws[1]]
    assert "Entry Date" not in blank_headers
    assert "Stage Started On" not in blank_headers
    assert blank_ws.cell(2, blank_headers.index("Current Stage") + 1).value is None
    assert blank_ws.cell(2, blank_headers.index("Candidate") + 1).protection.locked is False
    assert sheet_io.WORKBOOK_META_SHEET in blank_wb.sheetnames
    assert blank_wb[sheet_io.WORKBOOK_META_SHEET].sheet_state == "veryHidden"
    template_response = bot.app.test_client().get(
        "/api/channel/template", headers={"X-Auth": bot.PANEL_PASSWORD})
    assert template_response.status_code == 200
    assert pipeline_schema.filename_for(bot._kolkata_today().isoformat()) in (
        template_response.headers.get("Content-Disposition") or ""
    )
    client = bot.app.test_client()
    uploaded = client.post(
        "/api/channel/upload",
        headers={"X-Auth": bot.PANEL_PASSWORD},
        data={"by": "HR-01", "d": "2026-07-21",
              "file": (io.BytesIO(workbook), "channel.xlsx")},
        content_type="multipart/form-data",
    )
    assert uploaded.status_code == 200
    assert uploaded.get_json()["created"] == 1
    key = ("2026-07-21", "LinkedIn", "", 7)
    assert store.candidates[1]["status"] == "HR Screening"
    assert store.candidates[1]["filled_by"] == "HR-01"

    repeated_upload = client.post(
        "/api/channel/upload", headers={"X-Auth": bot.PANEL_PASSWORD},
        data={"by": "HR-01", "d": "2026-07-21",
              "file": (io.BytesIO(workbook), "channel.xlsx")},
        content_type="multipart/form-data")
    assert repeated_upload.status_code == 200
    assert repeated_upload.get_json()["created"] == 0
    assert repeated_upload.get_json()["updated"] == 1
    assert len(store.candidates) == 1

    stale_workbook = sheet_io._attach_workbook_metadata(
        sheet_io.build_xlsx(
            sheet_io.pipeline_columns(["Sales"]),
            prefill_rows=[{"name": "Stale", "channel": "LinkedIn", "job": "Sales"}],
            sheet_title=pipeline_schema.PIPELINE_TABLE_NAME,
        ),
        "2026-07-20", workbook_key,
    )
    stale_upload = client.post(
        "/api/channel/upload", headers={"X-Auth": bot.PANEL_PASSWORD},
        data={"by": "HR-01", "file": (io.BytesIO(stale_workbook), "stale.xlsx")},
        content_type="multipart/form-data",
    )
    assert stale_upload.status_code == 400
    assert "generated on 2026-07-20" in stale_upload.get_json()["error"]

    unsigned_upload = client.post(
        "/api/channel/upload", headers={"X-Auth": bot.PANEL_PASSWORD},
        data={"by": "HR-01", "file": (io.BytesIO(sheet_io.build_xlsx(
            sheet_io.pipeline_columns(["Sales"]), sheet_title=pipeline_schema.PIPELINE_TABLE_NAME
        )), "unsigned.xlsx")},
        content_type="multipart/form-data",
    )
    assert unsigned_upload.status_code == 400
    assert "no Nexus provenance record" in unsigned_upload.get_json()["error"]

    tampered_wb = load_workbook(io.BytesIO(blank_template))
    tampered_wb[sheet_io.WORKBOOK_META_SHEET]["B4"] = "2026-07-20"
    tampered_output = io.BytesIO()
    tampered_wb.save(tampered_output)
    tampered_upload = client.post(
        "/api/channel/upload", headers={"X-Auth": bot.PANEL_PASSWORD},
        data={"by": "HR-01", "file": (io.BytesIO(tampered_output.getvalue()), "tampered.xlsx")},
        content_type="multipart/form-data",
    )
    assert tampered_upload.status_code == 400
    assert "signature is invalid" in tampered_upload.get_json()["error"]

    rejected_csv = client.post(
        "/api/channel/upload",
        headers={"X-Auth": bot.PANEL_PASSWORD},
        data={"file": (io.BytesIO(b"a,b"), "channel.csv")},
        content_type="multipart/form-data",
    )
    assert rejected_csv.status_code == 400

    bot.lark_bitable.list_pipeline_records = lambda app, table: {"ok": True, "records": []}
    ensure_calls = []
    bot.lark_bitable.ensure_channel_base_schema = lambda app, pipeline, manual, **kwargs: (
        ensure_calls.append((app, pipeline, manual)) or {"ok": True}
    )
    bot.lark_bitable.prepare_canonical_manual_table = lambda app, manual, jobs, channels: {
        "ok": True, "changed": False, "table_id": manual,
    }
    bot.lark_bitable.delete_known_unsynced_test_rows = lambda app, pipeline: {
        "ok": True, "removed": []
    }
    bot.lark_bitable.verify_channel_base_schema = lambda app, pipeline, manual, **kwargs: {"ok": True}
    bot.lark_bitable.list_channel_records = lambda app, table: {"ok": True, "records": [{
        "record_id": "rec-1", "fields": {
            "日期": "2026-07-21", "招聘渠道": "LinkedIn", "关联职位": "Sales",
            "今日新增简历数": 11, "初筛通过数": 5,
            "已推荐面试数": 2, "已拒绝数": 3, "填写人": "HR-01",
        },
    }]}
    synced = client.post("/api/lark/pull", headers={"X-Auth": bot.PANEL_PASSWORD})
    assert synced.status_code == 200 and synced.get_json()["applied"] == 1
    assert store.rows[key]["new_resumes"] == 11

    missing_other_detail = client.post("/api/candidates", headers={"X-Auth": bot.PANEL_PASSWORD}, json={
        "apply_date": "2026-07-21", "stage_date": "2099-12-31", "name": "Ravi",
        "channel": "Other", "job_request_id": 7, "status": "Interview 1",
    })
    assert missing_other_detail.status_code == 422

    forbidden_other_detail = client.post(
        "/api/candidates", headers={"X-Auth": bot.PANEL_PASSWORD}, json={
            "name": "Ravi", "channel": "Naukri", "source_detail": "not allowed",
            "job_request_id": 7, "status": "New Lead",
        })
    assert forbidden_other_detail.status_code == 422

    direct_entry = client.post("/api/candidates", headers={"X-Auth": bot.PANEL_PASSWORD}, json={
        "apply_date": "2026-07-21", "stage_date": "2099-12-31", "name": "Ravi",
        "channel": "Naukri", "source_detail": "", "job_request_id": 7,
        "status": "Interview 1", "filled_by": "HR-01",
    })
    assert direct_entry.status_code == 200
    candidate_id = direct_entry.get_json()["id"]
    assert store.candidates[candidate_id]["status"] == "Interview 1"
    assert store.candidates[candidate_id]["stage_date"] != "2099-12-31"
    manager_row = bot._cand_json({
        **store.candidates[candidate_id],
        "stage_date": bot._kolkata_today().isoformat(),
    })
    assert manager_row["days_in_stage"] == 0

    rejected_without_reason = client.patch(
        "/api/candidates/%d" % candidate_id, headers={"X-Auth": bot.PANEL_PASSWORD}, json={
            "name": "Ravi", "channel": "Naukri", "job_request_id": 7,
            "status": "Rejected", "stage_date": "2026-07-22",
        })
    assert rejected_without_reason.status_code == 422

    rejected = client.patch(
        "/api/candidates/%d" % candidate_id, headers={"X-Auth": bot.PANEL_PASSWORD}, json={
            "name": "Ravi", "channel": "Naukri", "job_request_id": 7,
            "status": "Rejected", "stage_date": "2026-07-22",
            "rejection_reason": "Compensation mismatch", "filled_by": "HR-01",
        })
    assert rejected.status_code == 200
    assert store.candidates[candidate_id]["status"] == "Rejected"

    cards = []
    texts = []
    bot.send_card = lambda chat_id, card: cards.append((chat_id, card)) or "message-id"
    bot.send_text = lambda chat_id, text: texts.append((chat_id, text)) or "message-id"
    bot.handle_dm("admin-open-id", "chat-id", "/channel_sheet")
    assert "RECRUITMENT BOT" in texts[-1][1]
    bot.handle_dm("admin-open-id", "chat-id", "/submit_channel_sheet")
    assert "RECRUITMENT BOT" in texts[-1][1]

    bot.APP_ID = "cli_task_bot_task999"
    bot.RECRUITMENT_BOT_NAME = "RECRUITMENT BOT"
    bot.lark_bitable.APP_ID = "cli_recruitment_bot_rec8888"
    lark_status = client.get("/api/lark/status", headers={"X-Auth": bot.PANEL_PASSWORD})
    assert lark_status.status_code == 200
    assert lark_status.get_json()["bot_name"] == "RECRUITMENT BOT"
    assert lark_status.get_json()["app_id_tail"] == "_rec8888"
    assert lark_status.get_json()["app_id_tail"] != bot.APP_ID[-8:]
    assert lark_status.get_json()["url"] == (
        "https://example.test/base/channel?table=pipeline-table"
    )

    status = client.get(
        "/api/integration/v1/channel/status",
        headers={"Authorization": "Bearer " + "w" * 32},
    )
    assert status.status_code == 401
    bot.NEXUS_TALENT_WORKER_TOKEN = "w" * 32
    status = client.get(
        "/api/integration/v1/channel/status",
        headers={"Authorization": "Bearer " + "w" * 32},
    )
    assert status.status_code == 200 and status.get_json()["configured"] is True
    submitted = client.post(
        "/api/integration/v1/channel/submit",
        headers={"Authorization": "Bearer " + "w" * 32},
    )
    assert submitted.status_code == 200 and submitted.get_json()["applied"] == 1

    # A stale marker must not bypass real Lark metadata verification.
    store.settings["lark_channel_schema_ensured"] = pipeline_schema.SCHEMA_VERSION
    bot.lark_bitable.verify_channel_base_schema = lambda app, pipeline, manual, **kwargs: {
        "ok": False, "errors": ["Entry Date is still editable text"]
    }
    ensured = client.post(
        "/api/lark/ensure-schema", headers={"X-Auth": bot.PANEL_PASSWORD})
    assert ensured.status_code == 200
    assert ensure_calls[-1] == ("app-token", "pipeline-table", "manual-table")
    assert store.settings["lark_channel_schema_ensured"] == pipeline_schema.SCHEMA_VERSION
    bot.lark_bitable.verify_channel_base_schema = lambda app, pipeline, manual, **kwargs: {"ok": True}
    assert client.post("/api/lark/cleanup-default-table",
                       headers={"X-Auth": bot.PANEL_PASSWORD}).status_code == 404

    panel = Path("panel.html").read_text(encoding="utf-8")
    assert "清理空白默认页签" not in panel
    assert "mcStageDate" not in panel

    blocked_reset = client.post(
        "/api/lark/reconnect", headers={"X-Auth": bot.PANEL_PASSWORD},
        json={"confirmation": "RESET_UNSYNCED_CHANNEL_LINK"},
    )
    assert blocked_reset.status_code == 409
    store.settings["lark_channel_last_sync"] = ""
    repaired = client.post(
        "/api/lark/reconnect", headers={"X-Auth": bot.PANEL_PASSWORD},
        json={"confirmation": "RESET_UNSYNCED_CHANNEL_LINK"},
    )
    assert repaired.status_code == 200
    assert repaired.get_json()["candidate_data_changed"] is False
    assert store.settings["lark_channel_app_token"] == ""

    print("Channel website upload and Lark/Bot submission routes: PASSED")
    print("XLSX-only boundary, shared upsert service, source rules, direct-stage entry, and dual-Bot isolation: PASSED")


if __name__ == "__main__":
    main()
